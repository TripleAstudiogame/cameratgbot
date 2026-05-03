import os
import hashlib
import secrets
import re as _re
import io
from datetime import datetime, timedelta, timezone
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, Depends, Request, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, field_validator
import jwt
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from dotenv import load_dotenv

import db
import engine
import logging
logger = logging.getLogger("app")

load_dotenv()

# ── Configuration ──
PORT = int(os.getenv("PORT", "6565"))
ALLOWED_HOSTS = os.getenv("ALLOWED_HOSTS", "*")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "Amir")
ADMIN_PASSWORD_HASH = os.getenv("ADMIN_PASSWORD_HASH", "")
SECRET_KEY = os.getenv("SECRET_KEY", "")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24
MAX_PASSWORD_LENGTH = 128  # Prevent PBKDF2 DoS with huge passwords

if not SECRET_KEY:
    SECRET_KEY = secrets.token_hex(32)

import binascii

def _hash(pw, salt=None):
    if salt is None:
        salt = os.urandom(16)
    else:
        salt = binascii.unhexlify(salt)
    dk = hashlib.pbkdf2_hmac('sha256', pw.encode('utf-8'), salt, 100000)
    return binascii.hexlify(salt).decode('utf-8') + ':' + binascii.hexlify(dk).decode('utf-8')

def verify_pw(pw, hashed):
    if ':' not in hashed:
        return hashlib.sha256(pw.encode('utf-8')).hexdigest() == hashed
    salt, _ = hashed.split(':', 1)
    return _hash(pw, salt) == hashed

if not ADMIN_PASSWORD_HASH:
    ADMIN_PASSWORD_HASH = _hash("Amir")

@asynccontextmanager
async def lifespan(app):
    from logging.handlers import RotatingFileHandler
    log_handler = RotatingFileHandler("bot.log", maxBytes=5*1024*1024, backupCount=3, encoding="utf-8")
    logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s',
                        handlers=[log_handler, logging.StreamHandler()])
    logging.info("Starting Report Camera Engine...")
    logging.info(f"Port: {PORT} | Hosts: {ALLOWED_HOSTS}")
    
    # Initialize Admin User if not exists
    admin = db.get_user_by_username(ADMIN_USERNAME)
    import sqlite3
    conn = sqlite3.connect(db.DB_FILE)
    if not admin:
        conn.execute(
            "INSERT INTO users (username, password_hash, name, role) VALUES (?, ?, ?, 'admin')",
            (ADMIN_USERNAME, ADMIN_PASSWORD_HASH, "Administrator")
        )
    else:
        # Always sync the admin password and role from the current .env / defaults
        conn.execute(
            "UPDATE users SET password_hash = ?, role = 'admin' WHERE username = ?",
            (ADMIN_PASSWORD_HASH, admin['username'])
        )
    conn.commit()
    conn.close()

    # Write credentials to a text file for easy debugging
    try:
        with open("credentials.txt", "w", encoding="utf-8") as f:
            f.write(f"Логин: {ADMIN_USERNAME}\n")
            if os.getenv("ADMIN_PASSWORD_HASH"):
                f.write("Пароль: [Тот, чей хэш указан в файле .env]\n")
            else:
                f.write("Пароль: Amir\n")
    except Exception:
        pass

    engine.start_engine()
    yield

limiter = Limiter(key_func=get_remote_address)

# Disable Swagger/Redoc in production — prevents API schema leakage
app = FastAPI(title="Report Camera NVR SaaS", lifespan=lifespan,
              docs_url=None, redoc_url=None, openapi_url=None)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

# Dynamic CORS based on ALLOWED_HOSTS
_cors_origins = []
if ALLOWED_HOSTS == "*":
    _cors_origins = ["*"]
else:
    for h in ALLOWED_HOSTS.split(","):
        h = h.strip()
        if h:
            _cors_origins.append(f"http://{h}:{PORT}")
            _cors_origins.append(f"https://{h}")
    _cors_origins += [f"http://localhost:{PORT}", f"http://127.0.0.1:{PORT}"]

app.add_middleware(CORSMiddleware, allow_origins=_cors_origins,
                   allow_credentials=True, allow_methods=["GET","POST","PUT","PATCH","DELETE"], allow_headers=["*"])

@app.middleware("http")
async def security_headers(request: Request, call_next):
    r = await call_next(request)
    r.headers["X-Content-Type-Options"] = "nosniff"
    r.headers["X-Frame-Options"] = "DENY"
    r.headers["X-XSS-Protection"] = "1; mode=block"
    r.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    r.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    r.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://unpkg.com https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdn.jsdelivr.net; "
        "font-src 'self' data: https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data: https://unpkg.com https://cdn.jsdelivr.net; "
        "connect-src 'self' https://unpkg.com https://cdn.jsdelivr.net; "
        "frame-src 'self'; "
        "media-src 'self';"
    )
    # Cache control for development and security
    if request.url.path.startswith("/api/") or request.url.path == "/":
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        r.headers["Pragma"] = "no-cache"
    return r

os.makedirs("web", exist_ok=True)
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")

def create_token(data: dict):
    d = data.copy()
    d["exp"] = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    d["iss"] = "report-camera"
    d["aud"] = "report-camera-dashboard"
    return jwt.encode(d, SECRET_KEY, algorithm=ALGORITHM)

async def get_user(token: str = Depends(oauth2_scheme)):
    try:
        p = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM],
                       issuer="report-camera", audience="report-camera-dashboard")
        if not p.get("sub"): raise HTTPException(401)
        user = db.get_user_by_username(p["sub"])
        if not user: raise HTTPException(401)
        return user
    except jwt.PyJWTError: raise HTTPException(401, "Invalid token")

# Models
_BOT_TOKEN_RE = _re.compile(r'^\d+:.+$')

class OrgCreate(BaseModel):
    name: str; bot_token: str; mail_username: str; mail_password: str
    contact_name: str=""; contact_phone: str=""; contact_title: str=""; notes: str=""
    mail_check_interval: int = 0
    telegram_cooldown: int = 0
    cameras: dict = {}
    @field_validator('name','bot_token','mail_username','mail_password')
    @classmethod
    def not_empty(cls,v):
        if not v or not v.strip(): raise ValueError('Required')
        return v.strip()
    @field_validator('bot_token')
    @classmethod
    def valid_bot_token(cls,v):
        if not _BOT_TOKEN_RE.match(v.strip()): raise ValueError('Invalid bot token format')
        return v.strip()

class OrgUpdate(BaseModel):
    name: str; bot_token: str; mail_username: str; mail_password: str; is_active: bool
    contact_name: str=""; contact_phone: str=""; contact_title: str=""; notes: str=""
    mail_check_interval: int = 0
    telegram_cooldown: int = 0
    cameras: dict = {}
    @field_validator('name','mail_username')
    @classmethod
    def not_empty(cls,v):
        if not v or not v.strip(): raise ValueError('Required')
        return v.strip()

class PasswordChange(BaseModel):
    current_password: str; new_password: str
    @field_validator('new_password')
    @classmethod
    def password_limits(cls, v):
        if len(v) < 4: raise ValueError('Password must be at least 4 characters')
        if len(v) > MAX_PASSWORD_LENGTH: raise ValueError(f'Password must be at most {MAX_PASSWORD_LENGTH} characters')
        return v

class UserCreate(BaseModel):
    username: str; password: str
    name: str = ""; phone: str = ""
    @field_validator('username','password')
    @classmethod
    def not_empty(cls,v):
        if not v or not v.strip(): raise ValueError('Required')
        return v.strip()

class SettingsUpdate(BaseModel):
    mail_check_interval: str = "3"
    telegram_cooldown: str = "10"
    default_subscription_days: str = "365"

# ── Auth ──
@app.post("/api/login")
@limiter.limit("5/minute")
async def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends()):
    user = db.get_user_by_username(form_data.username)
    if user and verify_pw(form_data.password, user["password_hash"]):
        return {
            "access_token": create_token({"sub": user["username"], "role": user["role"], "id": user["id"], "pw_prefix": user["password_hash"][:8]}),
            "token_type": "bearer",
            "role": user["role"]
        }
    raise HTTPException(400, "Incorrect credentials")

# ── Users Management ──
@app.get("/api/users")
def list_users(u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    return db.get_users()

@app.post("/api/users")
def create_user(user: UserCreate, u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    try:
        new_id = db.add_user(user.username, _hash(user.password), user.name, user.phone)
        if not new_id:
            logger.warning(f"User creation failed: username {user.username} already exists")
            raise HTTPException(400, "Пользователь с таким логином уже существует")
        logger.info(f"User created: {user.username} (ID: {new_id})")
        return {"id": new_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error creating user: {e}")
        raise HTTPException(500, f"Internal Server Error: {str(e)}")

@app.delete("/api/users/{uid}")
def delete_user(uid: int, u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    db.delete_user(uid)
    return {"ok": True}

# ── System Status ──
@app.get("/api/system/update-status")
def get_update_status(u: dict = Depends(get_user)):
    try:
        if not os.path.exists("update.log"):
            return {"status": "Ожидание первого запуска..."}
        with open("update.log", "r", encoding="utf-8") as f:
            lines = f.readlines()
            if not lines:
                return {"status": "Нет данных"}
            last_line = lines[-1].strip()
            return {"status": last_line}
    except Exception as e:
        return {"status": f"Ошибка чтения лога: {e}"}

# ── Organizations CRUD ──
def _mask_org(org: dict) -> dict:
    """Remove sensitive fields from org data sent to frontend."""
    o = dict(org)
    if o.get('bot_token'):
        t = o['bot_token']
        o['bot_token'] = t[:6] + '***' + t[-4:] if len(t) > 10 else '***'
    if o.get('mail_password'):
        o['mail_password'] = '••••••••'
    return o

@app.get("/api/organizations")
def list_orgs(user_id: int = None, u: dict = Depends(get_user)):
    role = u.get("role")
    uid = u.get("id")
    if role == "admin":
        orgs = db.get_organizations(user_id)
    else:
        orgs = db.get_organizations(uid)
    
    print(f"[DEBUG] User {u.get('username')} (ID: {uid}, Role: {role}) requested orgs. Found: {len(orgs)}")
    return [_mask_org(o) for o in orgs]

@app.post("/api/organizations")
def create_org(org: OrgCreate, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    nid = db.add_organization(org.model_dump(), user_id=u["id"])
    db.add_event(nid, 'system', 'Организация создана (пробные 3 дня)')
    background_tasks.add_task(engine.reload_engine)
    return {"id": nid}

@app.put("/api/organizations/{oid}")
def update_org(oid: int, org: OrgUpdate, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    o = db.get_organization(oid)
    if not o or (u["role"] != "admin" and o.get("user_id") != u["id"]): raise HTTPException(403)
    data = org.model_dump()
    # If the frontend sends masked values back, keep the originals from DB
    if '***' in data.get('bot_token', '') or data.get('bot_token', '').startswith('•'):
        data['bot_token'] = o['bot_token']
    if data.get('mail_password', '') == '••••••••' or '•' in data.get('mail_password', ''):
        data['mail_password'] = o['mail_password']
    db.update_organization(oid, data); background_tasks.add_task(engine.reload_engine)
    return {"ok": True}

@app.delete("/api/organizations/{oid}")
def delete_org(oid: int, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    o = db.get_organization(oid)
    if not o or (u["role"] != "admin" and o.get("user_id") != u["id"]): raise HTTPException(403)
    db.delete_organization(oid); background_tasks.add_task(engine.reload_engine)
    return {"ok": True}

@app.patch("/api/organizations/{oid}/toggle")
def toggle_org(oid: int, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    new = not org['is_active']
    import sqlite3; conn=sqlite3.connect('organizations.db')
    conn.execute('UPDATE organizations SET is_active=? WHERE id=?',(int(new),oid)); conn.commit(); conn.close()
    db.add_event(oid, 'system', f'Бот {"запущен" if new else "остановлен"}')
    background_tasks.add_task(engine.reload_engine)
    return {"is_active": new}

class ExtendRequest(BaseModel):
    days: int

@app.post("/api/organizations/{oid}/extend_subscription")
def extend_sub(oid: int, req: ExtendRequest, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403, "Только администратор может продлевать подписку")
    nd = db.extend_subscription(oid, req.days)
    if not nd: raise HTTPException(404)
    db.add_event(oid, 'system', f'Подписка продлена до {nd[:10]}')
    background_tasks.add_task(engine.reload_engine)
    return {"new_date": nd}

@app.delete("/api/organizations/{oid}/subscribers/{cid}")
def remove_sub(oid: int, cid: int, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    if db.remove_subscriber(oid, cid):
        db.add_event(oid, 'subscriber', f'Подписчик удалён: {cid}')
        return {"ok": True}
    raise HTTPException(404)

# ── Access Management ──
@app.get("/api/organizations/{oid}/access")
def get_access(oid: int, status: str = None, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    return db.get_access_requests(oid, status if status else None)

@app.post("/api/organizations/{oid}/access/{rid}/approve")
def approve_access(oid: int, rid: int, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    req = db.approve_access_request(rid)
    if not req: raise HTTPException(404)
    db.add_event(oid, 'subscriber', f'Доступ одобрен: {req["first_name"]} {req["last_name"]} ({req["phone"]})')
    # Notify user via bot
    engine.notify_user_access(oid, req['chat_id'], approved=True)
    return {"ok": True}

@app.post("/api/organizations/{oid}/access/{rid}/reject")
def reject_access(oid: int, rid: int, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    db.reject_access_request(rid)
    db.add_event(oid, 'system', f'Заявка отклонена: #{rid}')
    return {"ok": True}

@app.delete("/api/organizations/{oid}/access/{rid}")
def revoke_user_access(oid: int, rid: int, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    req = db.revoke_access(rid)
    if not req: raise HTTPException(404)
    db.add_event(oid, 'subscriber', f'Доступ отозван: {req["first_name"]} ({req["chat_id"]})')
    # Notify user via bot
    engine.notify_user_access(oid, req['chat_id'], approved=False)
    background_tasks.add_task(engine.reload_engine)
    return {"ok": True}

# ── Test Connection ──
@app.post("/api/organizations/{oid}/test")
def test_conn(oid: int, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    result = engine.test_connection(org)
    status = "✅ Всё работает" if result['imap_ok'] and result['bot_ok'] else "❌ Есть ошибки"
    db.add_event(oid, 'system', f'Тест: {status}')
    return result

# ── Health (Heartbeat) ──
@app.get("/api/health")
def health(u: str = Depends(get_user)):
    return engine.get_health()

# ── Events (Activity Log) ──
@app.get("/api/events/{oid}")
def org_events(oid: int, u: dict = Depends(get_user)):
    org = db.get_organization(oid)
    if not org or (u["role"] != "admin" and org.get("user_id") != u["id"]): raise HTTPException(403)
    return db.get_events(oid, 50)

@app.get("/api/events")
def all_events(u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    return db.get_all_events(100)

# ── Analytics ──
@app.get("/api/analytics/daily")
def daily_stats(u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    return db.get_event_stats(days=30)

@app.get("/api/analytics/cameras")
def camera_stats(u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    return db.get_camera_stats(days=30)

# ── Settings ──
@app.get("/api/settings")
def get_settings(u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    return db.get_all_settings()

@app.put("/api/settings")
def update_settings(s: SettingsUpdate, background_tasks: BackgroundTasks, u: dict = Depends(get_user)):
    if u["role"] != "admin": raise HTTPException(403)
    db.set_setting('mail_check_interval', s.mail_check_interval)
    db.set_setting('telegram_cooldown', s.telegram_cooldown)
    db.set_setting('default_subscription_days', s.default_subscription_days)
    background_tasks.add_task(engine.reload_engine)
    return {"ok": True}

@app.post("/api/settings/change-password")
def change_password(data: PasswordChange, u: dict = Depends(get_user)):
    if not verify_pw(data.current_password, u["password_hash"]):
        raise HTTPException(400, "Неверный текущий пароль")
    
    new_hash = _hash(data.new_password)
    db.update_user_password(u["id"], new_hash)
    
    # If the user is admin, also update the global var (fallback)
    if u["username"] == ADMIN_USERNAME:
        global ADMIN_PASSWORD_HASH
        ADMIN_PASSWORD_HASH = new_hash
        try:
            env_path = os.path.join(os.path.dirname(__file__), '.env')
            with open(env_path, 'r') as f: content = f.read()
            import re
            content = re.sub(r'ADMIN_PASSWORD_HASH=.*', f'ADMIN_PASSWORD_HASH={ADMIN_PASSWORD_HASH}', content)
            with open(env_path, 'w') as f: f.write(content)
        except: pass
    return {"ok": True}

# ── Export Excel ──
@app.get("/api/export")
def export_excel(user_id: int = None, type: str = None, u: dict = Depends(get_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Организации"
    
    headers = ["ID","Название","Почта","Статус","Подписчиков","Оплачено до","Контакт","Телефон","Должность","Заметки"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    if u["role"] == "admin":
        if type == "mine":
            orgs = db.get_organizations(u["id"])
        else:
            orgs = db.get_organizations(user_id)
    else:
        orgs = db.get_organizations(u["id"])

    for i, org in enumerate(orgs, 2):
        vals = [org['id'], org['name'], org['mail_username'],
                'Активен' if org['is_active'] else 'Остановлен',
                len(org['subscribers']),
                org.get('subscription_end_date','')[:10],
                org.get('contact_name',''), org.get('contact_phone',''),
                org.get('contact_title',''), org.get('notes','')]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
    
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 18
    
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"report_camera_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            headers={"Content-Disposition": f"attachment; filename={fname}"})

# ── Frontend ──
app.mount("/static", StaticFiles(directory="web"), name="static")

@app.get("/")
def dashboard(): return FileResponse("web/index.html")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=PORT, reload=False)
